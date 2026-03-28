#!/usr/bin/env python3
"""
gyrotest_parem_full.py

Parse gyro runs, integrate with trapezoid rule, save CSV summaries, and provide
an interactive plotter that uses Plotly if available, otherwise Matplotlib.
"""

import sys, re, math, csv
from pathlib import Path
import numpy as np
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# ---------- config ----------
DATA_FOLDER = "SERVO_DATA"
DEFAULT_INPUT = "/mnt/data/calib_data_20260224-2156.txt"
OUT_RUN_SUMMARY = "runs_summary.csv"
OUT_SAMPLES = "runs_samples.csv"
BIAS_STRING = "&GYRO_BIAS GX2.207271GY1.248071GZ0.380602"

# regexes (robust)
re_header = re.compile(r"[&]?\s*R\s*(?P<run>\d+)\s*S\s*(?P<set>\d+)\s*V\s*(?P<v>[-\d\.eE]+)", re.IGNORECASE)
re_header_alt = re.compile(r"[&]?\s*R(?P<run>\d+)S(?P<set>\d+)V(?P<v>[-\d\.eE]+)", re.IGNORECASE)
re_t_line = re.compile(r"[&]?T(?P<ts>\d+).*?GX(?P<gx>[-\d\.eE]+)GY(?P<gy>[-\d\.eE]+)GZ(?P<gz>[-\d\.eE]+)")
re_mode_start = re.compile(r"MODE=GYRO_START", re.IGNORECASE)
re_mode_end = re.compile(r"MODE=GYRO_END", re.IGNORECASE)
re_bias = re.compile(r"GX(?P<gx>[-\d\.eE]+).*?GY(?P<gy>[-\d\.eE]+).*?GZ(?P<gz>[-\d\.eE]+)")

def build_Tg(g_yz,g_zy,g_xz,g_zx,g_xy,g_yx):
    # same convention you used elsewhere
    return np.array([[1.0, -g_yz,  g_zy],
                     [ g_xz, 1.0, -g_zx],
                     [-g_xy, g_yx, 1.0]], dtype=float)

def parse_csv_floats(s):
    # Accept "a,b,c" or "a b c" and return list of floats
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    for sep in [",", " "]:
        parts = [p.strip() for p in s.split(sep) if p.strip() != ""]
        if len(parts) >= 1:
            try:
                vals = [float(p) for p in parts]
                return vals
            except Exception:
                continue
    # last attempt: try eval-style
    try:
        vals = [float(p) for p in re.split(r"[,\s]+", s) if p]
        return vals
    except Exception:
        return None

def parse_gyro_cal_arg(s):
    """
    Parse --gyro-cal value.
    Accepts comma/space-separated floats.
    Returns tuple (bias3 or None, scale3 or None, mis6 or None).
    Interpretations:
      - 3 values  -> bias (gx,gy,gz)
      - 6 values  -> misalign (g_yz,g_zy,g_xz,g_zx,g_xy,g_yx)
      - 12 values -> bias(3), scale(3), misalign(6)
    """
    vals = parse_csv_floats(s)
    if vals is None:
        return None, None, None
    n = len(vals)
    if n == 3:
        return tuple(vals), None, None
    if n == 6:
        return None, None, tuple(vals)
    if n == 12:
        bias = tuple(vals[0:3])
        scale = tuple(vals[3:6])
        mis = tuple(vals[6:12])
        return bias, scale, mis
    raise ValueError("Invalid --gyro-cal format: expected 3, 6 or 12 numeric values (bias / misalign / bias+scale+misalign).")


def parse_multiple_gyro_cal_sets(s):
    """
    Parse multiple calibration sets from one input string.
    Accept separators: semicolon or newline.
    Returns a list of calibration strings (or [None] if empty).
    """
    if s is None:
        return [None]
    raw = str(s).strip()
    if not raw:
        return [None]

    parts = [p.strip() for p in re.split(r"[;\n]+", raw) if p.strip()]
    if not parts:
        return [None]

    # Validate each set early for clear user feedback.
    for part in parts:
        parse_gyro_cal_arg(part)
    return parts


def parse_bias_from_string(s):
    m = re_bias.search(s)
    if not m:
        return (0.0,0.0,0.0)
    return (float(m.group("gx")), float(m.group("gy")), float(m.group("gz")))

def select_input_path(input_path_from_cli):
    # if CLI path provided
    if input_path_from_cli:
        p = Path(input_path_from_cli)
        if p.exists():
            return str(p)
        else:
            raise FileNotFoundError(f"Provided path not found: {input_path_from_cli}")

    # try default uploaded path
    p = Path(DEFAULT_INPUT)
    if p.exists():
        return str(p)

    # fallback to DATA_FOLDER listing
    folder = Path(DATA_FOLDER)
    if not folder.exists():
        raise FileNotFoundError(f"No input and DATA_FOLDER '{DATA_FOLDER}' not found.")
    files = sorted([f for f in folder.iterdir() if f.is_file() and f.suffix.lower()=='.txt'])
    if not files:
        raise FileNotFoundError(f"No .txt in {DATA_FOLDER}")
    print("Available files:")
    for i,f in enumerate(files):
        print(f" {i}: {f.name}")
    idx = input("Select file index (Enter for latest): ").strip()
    if idx == "":
        return str(files[-1])
    idx = int(idx)
    return str(files[idx])

def apply_gyro_calib_to_samples(samples, mis_params=None, scale_params=None, bias_params=None):
    """
    samples: list of (ts, gx, gy, gz)
    mis_params: 6-element list or None -> [g_yz,g_zy,g_xz,g_zx,g_xy,g_yx]
    scale_params: 3-element list or None -> [sx,sy,sz]
    bias_params: 3-element list or None -> [bx,by,bz]
    returns: new_samples list of (ts, gx_cal, gy_cal, gz_cal)
    """
    if not samples:
        return samples
    bias = np.array(bias_params, dtype=float) if bias_params is not None else np.zeros(3)
    scales = np.array(scale_params, dtype=float) if scale_params is not None else np.ones(3)
    if mis_params is not None:
        if len(mis_params) != 6:
            raise ValueError("misalignment requires 6 parameters (g_yz,g_zy,g_xz,g_zx,g_xy,g_yx)")
        Tg = build_Tg(*mis_params)
    else:
        Tg = np.eye(3)
    Kg = np.diag(scales)

    out = []
    for (ts, gx, gy, gz) in samples:
        v = np.array([gx, gy, gz], dtype=float)
        v_corr = Tg @ (Kg @ (v - bias))
        out.append((ts, float(v_corr[0]), float(v_corr[1]), float(v_corr[2])))
    return out

def integrate_run_trapezoid(samples, bias):
    if not samples:
        return None
    samples = sorted(samples, key=lambda s: s[0])
    ts = np.array([s[0] for s in samples], dtype=np.int64)
    gx = np.array([s[1] for s in samples], dtype=float)
    gy = np.array([s[2] for s in samples], dtype=float)
    gz = np.array([s[3] for s in samples], dtype=float)
    gx_u = gx - bias[0]; gy_u = gy - bias[1]; gz_u = gz - bias[2]
    t_s = ts.astype(np.float64) / 1e6
    if len(t_s) < 2:
        return {
            "duration_s": 0.0,
            "start_ts_us": int(ts[0]),
            "end_ts_us": int(ts[-1]),
            "N": len(ts),
            "final_ang_x_deg": 0.0,
            "final_ang_y_deg": 0.0,
            "final_ang_z_deg": 0.0,
            "total_ang_disp_deg": 0.0,
            "per_sample": [{"ts_us":int(ts[0]), "gx_u":float(gx_u[0]), "gy_u":float(gy_u[0]), "gz_u":float(gz_u[0]),
                            "angle_x_deg":0.0,"angle_y_deg":0.0,"angle_z_deg":0.0}]
        }
    dt = np.diff(t_s)
    ang_x = np.sum(0.5*(gx_u[1:]+gx_u[:-1])*dt)
    ang_y = np.sum(0.5*(gy_u[1:]+gy_u[:-1])*dt)
    ang_z = np.sum(0.5*(gz_u[1:]+gz_u[:-1])*dt)
    w_prev = np.sqrt(gx_u[:-1]**2 + gy_u[:-1]**2 + gz_u[:-1]**2)
    w_curr = np.sqrt(gx_u[1:]**2 + gy_u[1:]**2 + gz_u[1:]**2)
    total_disp = np.sum(0.5*(w_prev + w_curr)*dt)
    per_sample = []
    cumx = cumy = cumz = 0.0
    per_sample.append({"ts_us": int(ts[0]), "gx_u": float(gx_u[0]), "gy_u": float(gy_u[0]), "gz_u": float(gz_u[0]),
                       "angle_x_deg": cumx, "angle_y_deg": cumy, "angle_z_deg": cumz})
    for i in range(1, len(ts)):
        dti = float(t_s[i] - t_s[i-1])
        step_x = 0.5*(gx_u[i]+gx_u[i-1])*dti
        step_y = 0.5*(gy_u[i]+gy_u[i-1])*dti
        step_z = 0.5*(gz_u[i]+gz_u[i-1])*dti
        cumx += step_x; cumy += step_y; cumz += step_z
        per_sample.append({"ts_us": int(ts[i]), "gx_u": float(gx_u[i]), "gy_u": float(gy_u[i]), "gz_u": float(gz_u[i]),
                           "angle_x_deg": float(cumx), "angle_y_deg": float(cumy), "angle_z_deg": float(cumz)})
    return {
        "duration_s": float(t_s[-1] - t_s[0]),
        "start_ts_us": int(ts[0]),
        "end_ts_us": int(ts[-1]),
        "N": int(len(ts)),
        "final_ang_x_deg": float(ang_x),
        "final_ang_y_deg": float(ang_y),
        "final_ang_z_deg": float(ang_z),
        "total_ang_disp_deg": float(total_disp),
        "per_sample": per_sample
    }


def list_input_files():
    folder = Path(DATA_FOLDER)
    if not folder.exists():
        return []
    return sorted([path for path in folder.rglob("*.txt") if path.is_file()])


def parse_runs_from_file(path):
    runs = []
    pending_hdr = None
    current_run = None
    bias_from_string = parse_bias_from_string(BIAS_STRING)

    with open(path, "r", encoding="utf-8", errors="ignore") as fh:
        for ln in fh:
            line = ln.strip()
            if not line:
                continue
            if "&GYRO_BIAS" in line:
                bias_from_string = parse_bias_from_string(line)
                continue

            mh = re_header.search(line) or re_header_alt.search(line)
            if mh:
                pending_hdr = {
                    "run_idx": int(mh.group("run")),
                    "set_id": int(mh.group("set")),
                    "V_cmd": float(mh.group("v")),
                    "hdr_line": line,
                }
                if current_run is not None:
                    runs.append(current_run)
                    current_run = None
                continue

            if re_mode_start.search(line):
                if current_run is not None:
                    runs.append(current_run)
                current_run = {
                    "run_idx": pending_hdr["run_idx"] if pending_hdr else None,
                    "set_id": pending_hdr["set_id"] if pending_hdr else None,
                    "V_cmd": pending_hdr["V_cmd"] if pending_hdr else float("nan"),
                    "samples": [],
                }
                pending_hdr = None
                continue

            if re_mode_end.search(line):
                if current_run is not None:
                    runs.append(current_run)
                    current_run = None
                pending_hdr = None
                continue

            m = re_t_line.search(line)
            if m:
                ts = int(m.group("ts"))
                gx = float(m.group("gx"))
                gy = float(m.group("gy"))
                gz = float(m.group("gz"))
                if current_run is None and pending_hdr is not None:
                    current_run = {
                        "run_idx": pending_hdr["run_idx"],
                        "set_id": pending_hdr["set_id"],
                        "V_cmd": pending_hdr["V_cmd"],
                        "samples": [],
                    }
                    pending_hdr = None
                if current_run is None:
                    current_run = {"run_idx": None, "set_id": None, "V_cmd": float("nan"), "samples": []}
                current_run["samples"].append((ts, gx, gy, gz))

    if current_run is not None:
        runs.append(current_run)

    return runs, bias_from_string


def process_gyro_file(path, gyro_cal_text=None):
    if gyro_cal_text:
        cli_bias, cli_scale, cli_mis = parse_gyro_cal_arg(gyro_cal_text)
    else:
        cli_bias = cli_scale = cli_mis = None

    runs, bias_from_string = parse_runs_from_file(path)
    raw_runs = [
        {
            "run_idx": run.get("run_idx"),
            "set_id": run.get("set_id"),
            "V_cmd": run.get("V_cmd", float("nan")),
            "samples": list(run.get("samples", [])),
        }
        for run in runs
    ]

    use_cli_cal = (cli_bias is not None) or (cli_scale is not None) or (cli_mis is not None)
    if use_cli_cal:
        for run in runs:
            run["samples"] = apply_gyro_calib_to_samples(
                run["samples"],
                mis_params=cli_mis,
                scale_params=cli_scale,
                bias_params=cli_bias,
            )
        bias_for_integration = (0.0, 0.0, 0.0)
    else:
        bias_for_integration = bias_from_string

    raw_bias_for_integration = bias_from_string

    all_samples = []
    runs_summary = []
    for order_idx, (run, raw_run) in enumerate(zip(runs, raw_runs), start=1):
        integration = integrate_run_trapezoid(run["samples"], bias_for_integration)
        raw_integration = integrate_run_trapezoid(raw_run["samples"], raw_bias_for_integration)
        if integration is None:
            continue

        duration = integration["duration_s"]
        curr_V = run.get("V_cmd", float("nan"))
        raw_total_disp = raw_integration["total_ang_disp_deg"] if raw_integration is not None else float("nan")

        runs_summary.append({
            "order_idx": order_idx,
            "run_idx": run.get("run_idx"),
            "set_id": run.get("set_id"),
            "V_cmd": curr_V,
            "start_ts_us": integration["start_ts_us"],
            "end_ts_us": integration["end_ts_us"],
            "duration_s": integration["duration_s"],
            "final_ang_x_deg": integration["final_ang_x_deg"],
            "final_ang_y_deg": integration["final_ang_y_deg"],
            "final_ang_z_deg": integration["final_ang_z_deg"],
            "raw_total_ang_disp_deg": raw_total_disp,
            "calib_total_ang_disp_deg": integration["total_ang_disp_deg"],
            "N_samples": integration["N"],
        })

        for sample in integration["per_sample"]:
            all_samples.append({
                "order_idx": order_idx,
                "run_idx": run.get("run_idx"),
                "set_id": run.get("set_id"),
                "V_cmd": curr_V,
                "ts_us": sample["ts_us"],
                "gx_u": sample["gx_u"],
                "gy_u": sample["gy_u"],
                "gz_u": sample["gz_u"],
                "angle_x_deg": sample["angle_x_deg"],
                "angle_y_deg": sample["angle_y_deg"],
                "angle_z_deg": sample["angle_z_deg"],
            })

    return {
        "path": str(path),
        "runs_summary": runs_summary,
        "all_samples": all_samples,
        "bias_from_string": bias_from_string,
        "applied_calibration": {
            "bias": cli_bias,
            "scale": cli_scale,
            "misalignment": cli_mis,
        },
    }


def save_csv_results(runs_summary, all_samples, summary_path, samples_path):
    if runs_summary:
        with open(summary_path, "w", newline="", encoding="utf-8") as fo:
            writer = csv.DictWriter(fo, fieldnames=list(runs_summary[0].keys()))
            writer.writeheader()
            for row in runs_summary:
                writer.writerow(row)

    if all_samples:
        with open(samples_path, "w", newline="", encoding="utf-8") as fo:
            writer = csv.DictWriter(fo, fieldnames=list(all_samples[0].keys()))
            writer.writeheader()
            for row in all_samples:
                writer.writerow(row)

def interactive_plotter(all_samples, runs_summary):
    """
    Interactive plotting helper. Uses Plotly if available, otherwise Matplotlib.
    Call this from main() after all_samples and runs_summary are prepared.
    """
    # safe availability checks (no importlib.util)
    try:
        import plotly.graph_objects as go
        _have_plotly = True
    except Exception:
        go = None
        _have_plotly = False

    try:
        import mplcursors
        _have_mplcursors = True
    except Exception:
        _have_mplcursors = False

    try:
        import seaborn as sns
        sns.set_theme(context="notebook", style="darkgrid")
    except Exception:
        pass

    import matplotlib.pyplot as plt

    # build lookup
    samples_by_run = {}
    runs_derived = {}
    for s in all_samples:
        samples_by_run.setdefault(s["order_idx"], []).append(s)

    for order_idx, rows in samples_by_run.items():
        rows_sorted = sorted(rows, key=lambda r: r["ts_us"])
        ts = np.array([r["ts_us"] for r in rows_sorted], dtype=np.int64)
        t_rel = (ts - ts[0]) / 1e6
        angle_x = np.array([r["angle_x_deg"] for r in rows_sorted], dtype=float)
        angle_y = np.array([r["angle_y_deg"] for r in rows_sorted], dtype=float)
        angle_z = np.array([r["angle_z_deg"] for r in rows_sorted], dtype=float)
        gx = np.array([r["gx_u"] for r in rows_sorted], dtype=float)
        gy = np.array([r["gy_u"] for r in rows_sorted], dtype=float)
        gz = np.array([r["gz_u"] for r in rows_sorted], dtype=float)
        t_s = ts.astype(np.float64) / 1e6
        if len(t_s) >= 2:
            dt = np.diff(t_s)
            w_prev = np.sqrt(gx[:-1]**2 + gy[:-1]**2 + gz[:-1]**2)
            w_curr = np.sqrt(gx[1:]**2 + gy[1:]**2 + gz[1:]**2)
            inc = 0.5 * (w_prev + w_curr) * dt
            cum_abs = np.concatenate(([0.0], np.cumsum(inc)))
        else:
            cum_abs = np.array([0.0])

        runs_derived[order_idx] = {
            "t_rel": t_rel,
            "angle_x": angle_x,
            "angle_y": angle_y,
            "angle_z": angle_z,
            "total_abs": cum_abs,
            "V_cmd": rows_sorted[0].get("V_cmd", float("nan")) if rows_sorted else float("nan")
        }

    # plotting helpers
    def _plotly_single(order_idx, axis="x", flip=False, abs_total=False, save_html=None):
        d = runs_derived.get(order_idx)
        if d is None:
            print("Run", order_idx, "not found.")
            return
        t = d["t_rel"]
        y = d["total_abs"] if abs_total else {"x": d["angle_x"], "y": d["angle_y"], "z": d["angle_z"]}.get(axis, d["angle_x"])
        if flip: y = -y
        meta = next((rr for rr in runs_summary if rr["order_idx"] == order_idx), None)
        title = f"Run order {order_idx}"
        if meta:
            title += f" (run_idx={meta.get('run_idx')}, V_cmd={meta.get('V_cmd')})"
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=t, y=y, mode="lines+markers" if len(t) < 300 else "lines", name="angle"))
        fig.update_layout(title=title, xaxis_title="Time since run start (s)", yaxis_title="Angle (deg)", template="plotly_white")
        if save_html:
            fig.write_html(save_html, include_plotlyjs="cdn")
            print("Saved interactive HTML to", save_html)
        fig.show()

    def _plotly_all(axis="x", flip=False, abs_total=False, save_html=None):
        fig = go.Figure()
        for order_idx in sorted(runs_derived.keys()):
            d = runs_derived[order_idx]
            t = d["t_rel"]
            y = d["total_abs"] if abs_total else {"x": d["angle_x"], "y": d["angle_y"], "z": d["angle_z"]}.get(axis, d["angle_x"])
            if flip: y = -y
            fig.add_trace(go.Scatter(x=t, y=y, mode="lines", name=f"run {order_idx} (V={d.get('V_cmd')})", line=dict(width=1)))
        title = "All runs: " + ("total_abs" if abs_total else f"angle_{axis}")
        if flip: title += " (flipped)"
        fig.update_layout(title=title, xaxis_title="Time since run start (s)", yaxis_title="Angle (deg)", template="plotly_white",
                          legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0))
        if save_html:
            fig.write_html(save_html, include_plotlyjs="cdn")
            print("Saved interactive HTML to", save_html)
        fig.show()

    def _mpl_single(order_idx, axis="x", flip=False, abs_total=False):
        d = runs_derived.get(order_idx)
        if d is None:
            print("Run", order_idx, "not found.")
            return
        t = d["t_rel"]
        y = d["total_abs"] if abs_total else {"x": d["angle_x"], "y": d["angle_y"], "z": d["angle_z"]}.get(axis, d["angle_x"])
        if flip: y = -y
        meta = next((rr for rr in runs_summary if rr["order_idx"] == order_idx), None)
        title = f"Run order {order_idx}"
        if meta:
            title += f" (run_idx={meta.get('run_idx')}, V_cmd={meta.get('V_cmd')})"
        plt.rcParams.update({"figure.dpi": 140, "lines.linewidth": 1.4, "font.size": 11})
        fig, ax = plt.subplots(figsize=(8,4))
        ax.plot(t, y, marker='o' if len(t) < 300 else None, markersize=4)
        ax.set_xlabel("Time since run start (s)")
        ax.set_ylabel("Angle (deg)")
        ax.set_title(title)
        ax.grid(True)
        plt.tight_layout()
        if _have_mplcursors:
            try:
                import mplcursors
                mplcursors.cursor(ax.get_lines(), hover=True)
            except Exception:
                pass
        plt.show()

    def _mpl_all(axis="x", flip=False, abs_total=False):
        plt.rcParams.update({"figure.dpi": 120, "lines.linewidth": 1.0, "font.size": 10})
        fig, ax = plt.subplots(figsize=(10,6))
        for order_idx in sorted(runs_derived.keys()):
            d = runs_derived[order_idx]
            t = d["t_rel"]
            y = d["total_abs"] if abs_total else {"x": d["angle_x"], "y": d["angle_y"], "z": d["angle_z"]}.get(axis, d["angle_x"])
            if flip: y = -y
            ax.plot(t, y, label=f"run {order_idx} (V={d.get('V_cmd')})", linewidth=0.8)
        title = "All runs: " + ("total_abs" if abs_total else f"angle_{axis}")
        if flip: title += " (flipped)"
        ax.set_xlabel("Time since run start (s)")
        ax.set_ylabel("Angle (deg)")
        ax.set_title(title)
        ax.legend(ncol=2, fontsize=8)
        ax.grid(True)
        plt.tight_layout()
        if _have_mplcursors:
            try:
                import mplcursors
                mplcursors.cursor(ax.get_lines(), hover=True)
            except Exception:
                pass
        plt.show()

    # interactive loop
    print("\nInteractive plotter ready. Commands:")
    print(" - Enter a run order number to plot that run (e.g. 1)")
    print(" - Modifiers (comma separated):")
    print("     flip      -> flip sign of plotted values")
    print("     abs       -> plot total absolute angle instead of axis")
    print("     axis=y    -> choose axis x/y/z (default x)")
    print("     html=PATH -> (plotly only) save interactive HTML to PATH and open it")
    print("   Examples: '3,abs'   '5,flip'   '7,axis=y,flip'  '2,html=run2.html'")
    print(" - Enter 'all' (or 'all,abs') to plot all runs")
    print(" - Enter 'list' to print available runs and V_cmd")
    print(" - Press Enter (empty) to exit")

    while True:
        cmd = input("\nPlot> ").strip()
        if cmd == "":
            print("Exiting plotter.")
            break
        parts = [p.strip() for p in cmd.split(",") if p.strip() != ""]
        if not parts:
            continue
        base = parts[0].lower()
        modifiers = parts[1:]
        flip_flag = any(m.lower() == "flip" for m in modifiers)
        abs_flag = any(m.lower() == "abs" for m in modifiers)
        axis_flag = "x"
        save_html = None
        for m in modifiers:
            ml = m.lower()
            if ml.startswith("axis="):
                axis_flag = ml.split("=",1)[1] or "x"
                axis_flag = axis_flag[0].lower()
            if ml.startswith("html="):
                save_html = m.split("=",1)[1]

        if base == "all":
            if _have_plotly and go is not None:
                _plotly_all(axis=axis_flag, flip=flip_flag, abs_total=abs_flag, save_html=save_html)
            else:
                _mpl_all(axis=axis_flag, flip=flip_flag, abs_total=abs_flag)
            continue
        if base == "list":
            for rr in runs_summary:
                print(f"order {rr['order_idx']}: run_idx={rr['run_idx']} set={rr['set_id']} V_cmd={rr['V_cmd']} dur={rr['duration_s']:.3f}s")
            continue
        try:
            idx = int(base)
            if _have_plotly and go is not None:
                _plotly_single(idx, axis=axis_flag, flip=flip_flag, abs_total=abs_flag, save_html=save_html)
            else:
                _mpl_single(idx, axis=axis_flag, flip=flip_flag, abs_total=abs_flag)
        except ValueError:
            print("Unrecognized command. Use a run number, 'all', 'list', or Enter to quit.")
            print("Modifiers: ',flip' to invert sign, ',abs' for total absolute angle, ',axis=y' to change axis, 'html=PATH' to save plotly HTML.")

class GyroTestGUI:
    def __init__(self, master):
        self.master = master
        master.title("Gyro Test Analyzer")
        master.geometry("1520x860")
        master.minsize(1320, 760)

        self.file_var = tk.StringVar()
        self.gyro_cal_var = tk.StringVar()
        self.status_var = tk.StringVar(value="Select a test file and run analysis.")
        self.last_result = None
        self.file_map = {}

        top = ttk.Frame(master, padding=10)
        top.pack(fill="x")
        top.columnconfigure(1, weight=1)

        ttk.Label(top, text="Test file:").grid(row=0, column=0, sticky="w")
        self.file_combo = ttk.Combobox(top, textvariable=self.file_var, state="readonly")
        self.file_combo.grid(row=0, column=1, sticky="ew", padx=(6, 6))
        ttk.Button(top, text="Refresh", command=self.refresh_file_list).grid(row=0, column=2, padx=4)

        ttk.Label(top, text="Gyro params (3/6/12 values):").grid(row=1, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(top, textvariable=self.gyro_cal_var).grid(row=1, column=1, sticky="ew", padx=(6, 6), pady=(8, 0))
        ttk.Label(top, text="3=bias, 6=misalignment, 12=bias scale misalignment; use ';' or newline for multiple sets").grid(row=1, column=2, sticky="w", pady=(8, 0))

        buttons = ttk.Frame(master, padding=(10, 0, 10, 6))
        buttons.pack(fill="x")
        self.run_button = ttk.Button(buttons, text="Run analysis", command=self.run_analysis)
        self.run_button.pack(side="left", padx=(0, 6))
        self.save_button = ttk.Button(buttons, text="Save XLSX", command=self.save_xlsx)
        self.save_button.pack(side="left", padx=6)
        self.quit_button = tk.Button(buttons, text="Quit", command=master.quit, bg="#c62828", fg="white", activebackground="#8e0000", activeforeground="white", padx=12)
        self.quit_button.pack(side="right")

        ttk.Label(master, textvariable=self.status_var, padding=(10, 0, 10, 8)).pack(anchor="w")

        columns = (
            "order_idx", "run_idx", "speed_deg_s", "duration_s",
            "final_ang_x_deg", "final_ang_y_deg", "final_ang_z_deg",
            "raw_total_ang_disp_deg", "calib_total_ang_disp_deg", "N_samples"
        )
        table_frame = ttk.LabelFrame(master, text="Per-run results", padding=(10, 6, 10, 10))
        table_frame.pack(fill="both", expand=True, padx=10, pady=(0, 8))
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings")
        headings = {
            "order_idx": "Order",
            "run_idx": "Run",
            "speed_deg_s": "Speed (deg/s)",
            "duration_s": "Duration (s)",
            "final_ang_x_deg": "Final X",
            "final_ang_y_deg": "Final Y",
            "final_ang_z_deg": "Final Z",
            "raw_total_ang_disp_deg": "Raw total disp",
            "calib_total_ang_disp_deg": "Calib total disp",
            "N_samples": "Samples",
        }
        widths = {
            "order_idx": 60,
            "run_idx": 60,
            "speed_deg_s": 100,
            "duration_s": 95,
            "final_ang_x_deg": 95,
            "final_ang_y_deg": 95,
            "final_ang_z_deg": 95,
            "raw_total_ang_disp_deg": 120,
            "calib_total_ang_disp_deg": 120,
            "N_samples": 80,
        }
        for column in columns:
            self.tree.heading(column, text=headings[column])
            self.tree.column(column, width=widths[column], anchor="center")
        y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        x_scroll = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        avg_frame = ttk.LabelFrame(master, text="Average angle by speed", padding=(10, 6, 10, 10))
        avg_frame.pack(fill="x", expand=False, padx=10, pady=(0, 10))
        avg_columns = ("speed_deg_s", "avg_raw_total_deg", "avg_calib_total_deg", "runs")
        self.avg_tree = ttk.Treeview(avg_frame, columns=avg_columns, show="headings", height=6)
        avg_headings = {
            "speed_deg_s": "Speed (deg/s)",
            "avg_raw_total_deg": "Avg raw angle (deg)",
            "avg_calib_total_deg": "Avg calib angle (deg)",
            "runs": "Runs",
        }
        avg_widths = {
            "speed_deg_s": 120,
            "avg_raw_total_deg": 180,
            "avg_calib_total_deg": 180,
            "runs": 80,
        }
        for column in avg_columns:
            self.avg_tree.heading(column, text=avg_headings[column])
            self.avg_tree.column(column, width=avg_widths[column], anchor="center")
        avg_scroll = ttk.Scrollbar(avg_frame, orient="vertical", command=self.avg_tree.yview)
        self.avg_tree.configure(yscrollcommand=avg_scroll.set)
        self.avg_tree.grid(row=0, column=0, sticky="ew")
        avg_scroll.grid(row=0, column=1, sticky="ns")
        avg_frame.columnconfigure(0, weight=1)

        multi_avg_frame = ttk.LabelFrame(master, text="Average by speed across parameter sets", padding=(10, 6, 10, 10))
        multi_avg_frame.pack(fill="x", expand=False, padx=10, pady=(0, 10))
        multi_columns = ("speed_deg_s", "avg_calib_across_sets_deg", "sets_used", "per_set_calib_deg")
        self.multi_avg_tree = ttk.Treeview(multi_avg_frame, columns=multi_columns, show="headings", height=6)
        multi_headings = {
            "speed_deg_s": "Speed (deg/s)",
            "avg_calib_across_sets_deg": "Avg calib across sets (deg)",
            "sets_used": "Sets used",
            "per_set_calib_deg": "Per-set calibrated avg (deg)",
        }
        multi_widths = {
            "speed_deg_s": 120,
            "avg_calib_across_sets_deg": 210,
            "sets_used": 90,
            "per_set_calib_deg": 760,
        }
        for column in multi_columns:
            self.multi_avg_tree.heading(column, text=multi_headings[column])
            self.multi_avg_tree.column(column, width=multi_widths[column], anchor="center")
        multi_scroll = ttk.Scrollbar(multi_avg_frame, orient="vertical", command=self.multi_avg_tree.yview)
        self.multi_avg_tree.configure(yscrollcommand=multi_scroll.set)
        self.multi_avg_tree.grid(row=0, column=0, sticky="ew")
        multi_scroll.grid(row=0, column=1, sticky="ns")
        multi_avg_frame.columnconfigure(0, weight=1)

        self.refresh_file_list()

    def refresh_file_list(self):
        files = list_input_files()
        self.file_map = {str(path.relative_to(Path(DATA_FOLDER))): path for path in files}
        labels = list(self.file_map.keys())
        self.file_combo["values"] = labels
        if labels and (not self.file_var.get() or self.file_var.get() not in self.file_map):
            self.file_var.set(labels[-1])
        if not labels:
            self.file_var.set("")
            self.status_var.set(f"No .txt files found in {DATA_FOLDER}.")

    def _set_buttons_state(self, state):
        self.run_button.config(state=state)
        self.save_button.config(state=state)
        self.quit_button.config(state=state)

    def _clear_tree(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

    def _clear_avg_tree(self):
        for item in self.avg_tree.get_children():
            self.avg_tree.delete(item)

    def _clear_multi_avg_tree(self):
        for item in self.multi_avg_tree.get_children():
            self.multi_avg_tree.delete(item)

    @staticmethod
    def _compute_speed_average_rows(runs_summary):
        grouped = {}
        for row in runs_summary:
            speed = row.get("V_cmd")
            is_missing_speed = speed is None
            if not is_missing_speed:
                try:
                    is_missing_speed = bool(np.isnan(float(speed)))
                except Exception:
                    is_missing_speed = True

            key = "Unknown" if is_missing_speed else float(speed)
            grouped.setdefault(key, []).append(row)

        ordered_keys = sorted([k for k in grouped.keys() if k != "Unknown"])
        if "Unknown" in grouped:
            ordered_keys.append("Unknown")

        out = []
        for speed in ordered_keys:
            rows = grouped[speed]
            raw_vals = [float(r["raw_total_ang_disp_deg"]) for r in rows]
            calib_vals = [float(r["calib_total_ang_disp_deg"]) for r in rows]
            avg_raw = float(np.mean(raw_vals)) if raw_vals else float("nan")
            avg_calib = float(np.mean(calib_vals)) if calib_vals else float("nan")
            out.append({
                "speed": speed,
                "avg_raw": avg_raw,
                "avg_calib": avg_calib,
                "runs": len(rows),
            })
        return out

    @staticmethod
    def _compute_across_set_speed_rows(per_set_results):
        # per_set_results: list of {label, result}
        grouped = {}
        for idx, item in enumerate(per_set_results, start=1):
            label = item.get("label") or f"set_{idx}"
            rows = GyroTestGUI._compute_speed_average_rows(item["result"]["runs_summary"])
            for row in rows:
                speed = row["speed"]
                g = grouped.setdefault(speed, {"calib": [], "labels": [], "raw": []})
                g["calib"].append(float(row["avg_calib"]))
                g["raw"].append(float(row["avg_raw"]))
                g["labels"].append(label)

        ordered_keys = sorted([k for k in grouped.keys() if k != "Unknown"])
        if "Unknown" in grouped:
            ordered_keys.append("Unknown")

        out = []
        for speed in ordered_keys:
            g = grouped[speed]
            avg_calib_across = float(np.mean(g["calib"])) if g["calib"] else float("nan")
            per_set_desc = " | ".join([f"{lbl}:{val:.6f}" for lbl, val in zip(g["labels"], g["calib"])])
            out.append({
                "speed": speed,
                "avg_calib_across": avg_calib_across,
                "sets_used": len(g["calib"]),
                "per_set_desc": per_set_desc,
            })
        return out

    def run_analysis(self):
        selected = self.file_var.get().strip()
        if not selected:
            messagebox.showerror("Missing file", "Select a test file first.")
            return
        path = self.file_map.get(selected)
        if path is None:
            messagebox.showerror("Invalid file", "Selected file is no longer available.")
            return

        self._set_buttons_state("disabled")
        self.status_var.set("Running analysis...")
        self._clear_tree()
        self._clear_avg_tree()
        self._clear_multi_avg_tree()
        self.master.update_idletasks()

        try:
            cal_sets = parse_multiple_gyro_cal_sets(self.gyro_cal_var.get())
            per_set_results = []
            for idx, cal_text in enumerate(cal_sets, start=1):
                label = "default" if cal_text is None else f"set_{idx}"
                per_set_results.append({
                    "label": label,
                    "cal_text": cal_text,
                    "result": process_gyro_file(path, cal_text),
                })

            primary = per_set_results[0]["result"]
            self.last_result = {
                "primary": primary,
                "per_set_results": per_set_results,
            }

            for row in primary["runs_summary"]:
                values = (
                    row["order_idx"],
                    self._fmt_int(row["run_idx"]),
                    self._fmt_float(row.get("V_cmd")),
                    self._fmt_float(row["duration_s"]),
                    self._fmt_float(row["final_ang_x_deg"]),
                    self._fmt_float(row["final_ang_y_deg"]),
                    self._fmt_float(row["final_ang_z_deg"]),
                    self._fmt_float(row["raw_total_ang_disp_deg"]),
                    self._fmt_float(row["calib_total_ang_disp_deg"]),
                    row["N_samples"],
                )
                self.tree.insert("", tk.END, values=values)

            avg_rows = self._compute_speed_average_rows(primary["runs_summary"])
            for avg_row in avg_rows:
                speed = avg_row["speed"]
                self.avg_tree.insert(
                    "",
                    tk.END,
                    values=(
                        speed if isinstance(speed, str) else self._fmt_float(speed),
                        self._fmt_float(avg_row["avg_raw"]),
                        self._fmt_float(avg_row["avg_calib"]),
                        avg_row["runs"],
                    ),
                )

            if len(per_set_results) > 1:
                multi_rows = self._compute_across_set_speed_rows(per_set_results)
                for multi_row in multi_rows:
                    speed = multi_row["speed"]
                    self.multi_avg_tree.insert(
                        "",
                        tk.END,
                        values=(
                            speed if isinstance(speed, str) else self._fmt_float(speed),
                            self._fmt_float(multi_row["avg_calib_across"]),
                            multi_row["sets_used"],
                            multi_row["per_set_desc"],
                        ),
                    )

            self.status_var.set(f"Processed {len(primary['runs_summary'])} runs from {selected} using {len(per_set_results)} parameter set(s)")
        except Exception as exc:
            self.last_result = None
            messagebox.showerror("Processing error", str(exc))
            self.status_var.set("Processing failed")
        finally:
            self._set_buttons_state("normal")

    def save_xlsx(self):
        if not self.last_result:
            messagebox.showinfo("No results", "Run analysis first.")
            return

        out_path = filedialog.asksaveasfilename(
            title="Save analysis workbook",
            defaultextension=".xlsx",
            initialfile="gyro_analysis.xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
        )
        if not out_path:
            return

        try:
            from openpyxl import Workbook

            primary = self.last_result["primary"]
            per_set_results = self.last_result.get("per_set_results", [{"label": "default", "result": primary}])

            runs_summary = primary["runs_summary"]
            avg_rows = self._compute_speed_average_rows(runs_summary)
            multi_rows = self._compute_across_set_speed_rows(per_set_results) if len(per_set_results) > 1 else []

            wb = Workbook()

            ws_runs = wb.active
            ws_runs.title = "per_run_results"
            run_headers = [
                "order_idx",
                "run_idx",
                "speed_deg_s",
                "duration_s",
                "final_ang_x_deg",
                "final_ang_y_deg",
                "final_ang_z_deg",
                "raw_total_ang_disp_deg",
                "calib_total_ang_disp_deg",
                "N_samples",
            ]
            ws_runs.append(run_headers)
            for row in runs_summary:
                ws_runs.append([
                    row.get("order_idx"),
                    row.get("run_idx"),
                    row.get("V_cmd"),
                    row.get("duration_s"),
                    row.get("final_ang_x_deg"),
                    row.get("final_ang_y_deg"),
                    row.get("final_ang_z_deg"),
                    row.get("raw_total_ang_disp_deg"),
                    row.get("calib_total_ang_disp_deg"),
                    row.get("N_samples"),
                ])

            ws_avg = wb.create_sheet("avg_by_speed")
            ws_avg.append(["speed_deg_s", "avg_raw_total_deg", "avg_calib_total_deg", "runs"])
            for row in avg_rows:
                ws_avg.append([
                    None if isinstance(row["speed"], str) else row["speed"],
                    row["avg_raw"],
                    row["avg_calib"],
                    row["runs"],
                ])

            if multi_rows:
                ws_multi = wb.create_sheet("avg_across_sets")
                ws_multi.append(["speed_deg_s", "avg_calib_across_sets_deg", "sets_used", "per_set_calib_deg"])
                for row in multi_rows:
                    ws_multi.append([
                        None if isinstance(row["speed"], str) else row["speed"],
                        row["avg_calib_across"],
                        row["sets_used"],
                        row["per_set_desc"],
                    ])

            sheets_to_format = [ws_runs, ws_avg]
            if multi_rows:
                sheets_to_format.append(ws_multi)

            for ws in sheets_to_format:
                for data_row in ws.iter_rows(min_row=2):
                    for cell in data_row:
                        if isinstance(cell.value, float):
                            cell.number_format = "0.000000"

            wb.save(out_path)
            self.status_var.set(f"Saved workbook to {Path(out_path).name}")
        except ImportError:
            messagebox.showerror("Missing dependency", "openpyxl is required for XLSX export. Install it with: pip install openpyxl")
        except Exception as exc:
            messagebox.showerror("Save error", str(exc))

    @staticmethod
    def _fmt_float(value):
        if value is None:
            return ""
        if isinstance(value, float) and math.isnan(value):
            return ""
        return f"{float(value):.6f}"

    @staticmethod
    def _fmt_int(value):
        if value is None:
            return ""
        return str(value)


def start_gui():
    root = tk.Tk()
    GyroTestGUI(root)
    root.mainloop()


if __name__ == "__main__":
    start_gui()